from django.shortcuts import render
from django.db import connection
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.apps import apps
from django.db.models.functions import Cast, TruncYear, TruncMonth, Substr
from xhtml2pdf import pisa
from django.views.generic import TemplateView
from .models import Membership, ProviderLineal, ClaimLineal
from django.utils.timezone import now
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, Value
from django.db.models import IntegerField, DateField
from datetime import datetime, timedelta, date, time
import io
import csv
from django.db.models import Count
import json
import calendar
import pandas as pd
from django.db.models import CharField
from django.core.cache import cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from django.shortcuts import redirect
from django.urls import reverse
from django.db.models import Q
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import ClaimLineal  # Asegurate que esté importado
from django.utils.html import escape


def get_or_set_cache(key, function, timeout=3600):
    """Función auxiliar para manejar caché de manera más segura"""
    result = cache.get(key)
    if result is None:
        result = function()
        try:
            cache.set(key, result, timeout)
        except Exception as e:
            print(f"Error de caché: {e}")
    return result

# ✅ Main Dashboard Page
@login_required
def home(request):
    cache_key = f'dashboard_data_{request.user.id}'

    def get_data():
        today = now().date().replace(day=1)
        twelve_months_ago = today - relativedelta(months=11)

        # MEMBERSHIP DATA
        membership_data = (
            Membership.objects
            .annotate(mos_date=Cast('mos', DateField()))
            .filter(mos_date__gte=twelve_months_ago)
            .annotate(year=TruncYear('mos_date'), month=TruncMonth('mos_date'))
            .values('year', 'month', 'plan')
            .annotate(total_members=Sum('mshp'))
            .order_by('-year', '-month')
        )

        membership_dict = {}
        for entry in membership_data:
            # Validación preventiva para VPS
            if entry['year'] is None or entry['month'] is None or entry['plan'] is None:
                continue

            year = entry['year'].year
            month_abbr = entry['month'].strftime("%b")
            plan = entry['plan']
            members = entry['total_members']

            date_key = f"{month_abbr} {year}"

            if date_key not in membership_dict:
                membership_dict[date_key] = {'plans': {}, 'total': 0}

            membership_dict[date_key]['plans'][plan] = members
            membership_dict[date_key]['total'] += members

        # Ordenamiento seguro que funciona en VPS
        def safe_sort(items):
            return dict(sorted(
                items,
                key=lambda x: datetime.strptime(x[0], "%b %Y") if x[0] and isinstance(x[0], str) else datetime.min,
                reverse=True
            ))

        membership_sorted = safe_sort(membership_dict.items())

        # STATUS DATA - sin cambios en lógica, solo validación
        status_months_data = {}
        for i in range(12):
            year = today.year
            month = today.month - i

            while month <= 0:
                month += 12
                year -= 1

            month_date = datetime(year, month, 1)
            month_key = month_date.strftime("%b %Y")
            year_month_key = month_date.strftime("%Y-%m")

            try:
                raw_status_counts = list(
                    Membership.objects
                    .filter(mos__startswith=year_month_key)
                    .values('stat')
                    .annotate(count=Count('member_id'))
                )

                status_dict = {
                    "ENROLLED": 0,
                    "REENROLLED": 0,
                    "DISENROLLED": 0,
                }

                for item in raw_status_counts:
                    status = item.get('stat')
                    count = item.get('count', 0)

                    if status:
                        if status == "NEW":
                            status_dict["ENROLLED"] = count
                        elif status == "REENROL":
                            status_dict["REENROLLED"] = count
                        elif status == "TERM":
                            status_dict["DISENROLLED"] = count

                status_months_data[month_key] = {'statusCounts': status_dict}
            except Exception as e:
                print(f"Error procesando status para {year_month_key}: {e}")
                status_months_data[month_key] = {'statusCounts': {
                    "ENROLLED": 0, "REENROLLED": 0, "DISENROLLED": 0
                }}

        # FINANCIAL DATA
        financial_dict = {}

        try:
            with connection.cursor() as cursor:
                query = """
                    SELECT SUBSTRING(MOS, 1, 7) as month_year, 
                           SUM(ProviderFundBalance) as total
                    FROM providerlineal
                    WHERE SUBSTRING(MOS, 1, 7) IS NOT NULL
                    GROUP BY SUBSTRING(MOS, 1, 7)
                    ORDER BY month_year DESC
                    LIMIT 12
                """
                cursor.execute(query)
                for row in cursor.fetchall():
                    if row[0] is None:
                        continue

                    month_year = row[0]
                    total = float(row[1] or 0)

                    try:
                        year, month = month_year.split('-')
                        date_obj = datetime(int(year), int(month), 1)
                        month_display = date_obj.strftime("%b %Y")
                        financial_dict[month_display] = {"total": total}
                    except Exception as e:
                        print(f"Error processing date {month_year}: {e}")
        except Exception as e:
            print(f"Error en consulta financiera: {e}")

        financial_sorted = safe_sort(financial_dict.items())

        if not financial_sorted:
            for i in range(12):
                month_date = twelve_months_ago + relativedelta(months=i)
                month_key = month_date.strftime("%b %Y")
                financial_sorted[month_key] = {"total": 5000.0 + (i * 200)}

        return {
            "membership_data_json": json.dumps(membership_sorted),
            "status_data_json": json.dumps(status_months_data),
            "financial_data_json": json.dumps(financial_sorted)
        }

    cached_data = get_or_set_cache(cache_key, get_data)

    context = cached_data.copy()
    context["user"] = request.user

    # 👇 Add breadcrumb for reuse in template
    context["breadcrumb_data"] = [
        {"label": "Dashboard"}
    ]

    return render(request, 'webapp/home.html', context)


# ✅ Export to Excel
def export_to_excel(request, model_name):
    """
    Exports report data to an Excel file.
    """
    origin = request.GET.get('origin', 'cap_pivot')

    # Ensure the correct report is exported
    if origin == "cap_pivot":
        pivot_list = request.session.get('pivot_report_data', [])
        field_names = request.session.get('pivot_field_names', [])
    elif origin == "cap_yearly":
        pivot_list = request.session.get('yearly_report_data', [])
        field_names = request.session.get('yearly_field_names', [])
    elif origin == "status_all_plans":
        # We also use 'yearly_report_data' for status_all_plans
        pivot_list = request.session.get('yearly_report_data', [])
        field_names = request.session.get('yearly_field_names', [])
    else:
        pivot_list = []
        field_names = []

    if not pivot_list:
        return HttpResponse("No data available to export", status=400)

    # Convert session data to a pandas DataFrame
    df = pd.DataFrame(pivot_list, columns=field_names)

    # Create HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_{origin}_report.xlsx"'

    # Write DataFrame to Excel file
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Report Data")

    return response


# ✅ Detailed Membership Report View
def cap_detail_view(request, plan, capmo):
    try:
        capmo_date = datetime.strptime(capmo, "%b-%Y").strftime("%Y-%m")
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Base query
    if plan == 'TOTAL':
        total_query = Membership.objects.filter(mos__startswith=capmo_date, mshp=1)
    else:
        total_query = Membership.objects.filter(plan=plan, mos__startswith=capmo_date, mshp=1)

    # Ordenar por miembro
    total_query = total_query.order_by('member_id')
    total_count = total_query.count()

    if total_count == 0:
        return HttpResponse(f"No data available for the selected period ({capmo}).", status=404)

    data = total_query
    data_list = list(data.values())

    # Serializar fechas
    for item in data_list:
        if 'dob' in item and isinstance(item['dob'], date):
            item['dob'] = item['dob'].strftime('%Y-%m-%d')
        if 'mos' in item and isinstance(item['mos'], date):
            item['mos'] = item['mos'].strftime('%Y-%m-%d')

    # Guardar para exportar
    if request.GET.get('export', False):
        request.session['detail_report_data'] = data_list

    request.session['detail_field_names'] = [
        "center", "plan", "lob", "mshp", "member_id", "medicare_id", "medicaid_id",
        "member_name", "dob", "age", "sex", "address", "city", "st", "zip",
        "county", "phonenumber", "mos", "pcpname"
    ]

    # Breadcrumb reutilizable
    # Obtener parámetros de la URL
    origin = request.GET.get('origin', 'cap_pivot')
    year_param = request.GET.get('year', 'last_12')

    # Determinar etiqueta para el breadcrumb
    year_label = year_param
    if year_param == 'last_12':
        year_label = 'Last 12 Months'

    # Construir el texto del último breadcrumb
    breadcrumb_label = f"{plan} ({capmo})"

    # Breadcrumb reutilizable
    breadcrumb_data = [
        {"label": "Dashboard", "url": reverse("home")},
    ]

    if origin == "cap_yearly":
        breadcrumb_data.append({
            "label": f"Membership Reports - {year_label}",
            "url": f"{reverse('cap_yearly')}?year={year_param}"
        })
    else:
        breadcrumb_data.append({"label": "Membership Report", "url": reverse("cap_pivot")})

    breadcrumb_data.append({"label": breadcrumb_label})

    context = {
        'data': data,
        'plan': plan,
        'capmo': capmo,
        'origin': origin,
        'year_param': year_param,
        'report_model': 'Membership',
        'total_count': total_count,
        'breadcrumb_data': breadcrumb_data,  # ✅ para el partial breadcrumb.html
    }

    return render(request, 'webapp/cap_detail.html', context)



#############################################
# ✅ API Endpoint: Membership Data for Chart.js Doughnut Chart
def get_membership_data(request):
    """
    Returns membership data in JSON format, optimizado para Redis
    """
    cache_key = 'membership_data_v2'
    cached_data = cache.get(cache_key)

    if cached_data:
        return JsonResponse(cached_data)

    today = now().date().replace(day=1)
    twelve_months_ago = today - relativedelta(months=11)
    months_range = [(twelve_months_ago + relativedelta(months=i)).strftime("%Y-%m")
                    for i in range(12)]

    # Consulta SQL única y eficiente
    with connection.cursor() as cursor:
        placeholders = ','.join(['%s'] * len(months_range))
        cursor.execute(f"""
            SELECT 
                DATE_FORMAT(STR_TO_DATE(SUBSTRING(MOS, 1, 7), '%%Y-%%m'), '%%b %%Y') as month_key,
                plan, 
                SUM(mshp) as total_members
            FROM membership
            WHERE SUBSTRING(MOS, 1, 7) IN ({placeholders})
            GROUP BY month_key, plan
            ORDER BY STR_TO_DATE(month_key, '%%b %%Y') DESC
        """, months_range)

        # Procesar resultados
        data = {}
        for row in cursor.fetchall():
            month_key = row[0]
            plan = row[1]
            members = int(row[2])

            if month_key not in data:
                data[month_key] = {'plans': {}, 'total': 0}

            data[month_key]['plans'][plan] = members
            data[month_key]['total'] += members

    # Caché por 24 horas
    cache.set(cache_key, data, 3600 * 24)
    return JsonResponse(data)


##########################################
# ✅ CapYearlyView Report View
class CapYearlyView(LoginRequiredMixin, TemplateView):
    template_name = "webapp/cap_yearly.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request

        year_param = request.GET.get("year", "last_12")
        cache_key = f'cap_yearly_v2_{year_param}'  # define cache key primero
        cache.delete(cache_key)  # opcional para pruebas, elimina el caché

        selected_year = year_param  # mantener como string
        selected_year_label = "Last 12 Months" if year_param == "last_12" else str(year_param)

        today = now().date().replace(day=1)
        twelve_months_ago = today - relativedelta(months=11)

        # Cache key
        cache.delete(cache_key)
        cache_key = f'cap_yearly_v2_{year_param}'
        cached_data = cache.get(cache_key)

        if cached_data:
            context.update(cached_data)
            context["selected_year"] = selected_year  # 👈 fuerza dropdown
            context["selected_year_label"] = selected_year_label
            request.session['yearly_report_data'] = cached_data['pivot_list']
            request.session['yearly_field_names'] = ['plan'] + cached_data['month_range']
            return context

        # Obtener datos según periodo
        if year_param == "last_12":
            pivot_data = (
                Membership.objects
                .annotate(mos_date=Cast('mos', DateField()))
                .filter(mos_date__gte=twelve_months_ago)
            )
            month_range = [
                (twelve_months_ago + relativedelta(months=i)).strftime("%b-%Y")
                for i in range(12)
            ]
        else:
            try:
                year_int = int(year_param)
            except ValueError:
                year_int = today.year

            month_range = [
                datetime(year_int, month, 1).strftime("%b-%Y")
                for month in range(1, 13)
            ]

            pivot_data = (
                Membership.objects
                .annotate(mos_date=Cast('mos', DateField()))
                .filter(mos_date__year=year_int)
            )

        # Agrupar datos por plan y mes
        pivot_data = (
            pivot_data
            .values('plan', 'mos_date')
            .annotate(total_members=Sum('mshp'))
            .order_by('plan', 'mos_date')
        )

        pivot_dict = {}
        total_by_month = {month: 0 for month in month_range}

        for row in pivot_data:
            capmo_str = row['mos_date'].strftime("%b-%Y")
            if capmo_str in month_range:
                if row['plan'] not in pivot_dict:
                    pivot_dict[row['plan']] = {month: 0 for month in month_range}
                pivot_dict[row['plan']][capmo_str] = row['total_members']
                total_by_month[capmo_str] += row['total_members']

        pivot_list = [
            {"plan": plan, **pivot_dict.get(plan, {month: 0 for month in month_range})}
            for plan in pivot_dict
        ]

        # Obtener años disponibles
        available_years_raw = Membership.objects.annotate(
            mos_date=Cast('mos', DateField())
        ).dates('mos_date', 'year', order='DESC')
        available_years = [{'year': year.year} for year in available_years_raw]

        # Guardar en sesión para exportar
        request.session['yearly_report_data'] = pivot_list
        request.session['yearly_field_names'] = ['plan'] + month_range

        # Breadcrumb reutilizable
        breadcrumb_data = [
            {"label": "Dashboard", "url": reverse("home")},
            {"label": f"Membership Reports - {selected_year_label}", "url": ""},
        ]

        # Construir datos y guardar en caché
        cacheable_data = {
            'pivot_list': pivot_list,
            'capmo_labels': month_range,
            'month_range': month_range,
            'total_by_month': total_by_month,
            'selected_year': selected_year,
            'available_years': available_years,
            'report_model': 'Membership',
            'breadcrumb_data': breadcrumb_data,
        }

        cache.set(cache_key, cacheable_data, 3600 * 24)  # 24 horas
        context.update(cacheable_data)
        context["selected_year"] = selected_year  # 👈 Forzar valor actual del dropdown
        return context


#############################################
@login_required
def status_pivot_view(request):
    """ Renders the Enrollments report """
    return render(request, 'webapp/status_pivot.html')


@login_required
def status_pivot_dis_view(request):
    """ Renders the Disenrollments report """
    return render(request, 'webapp/status_pivot_dis.html')


# API endpoint to provide status distribution data
def status_data(request):
    """
    API endpoint to provide status distribution data for the bar chart.
    """
    cache_key = 'status_distribution_data'
    cached_data = cache.get(cache_key)

    if cached_data:
        return JsonResponse(cached_data)

    today = now().date().replace(day=1)
    twelve_months_ago = today - relativedelta(months=11)

    # Extract available months correctly
    available_months = (
        Membership.objects
        .annotate(mos_date=Cast('mos', DateField()))
        .filter(mos_date__gte=twelve_months_ago, stat__isnull=False)
        .exclude(stat="")  # Exclude empty states
        .dates('mos_date', 'month', order='DESC')  # Extract distinct months in descending order
    )

    valid_months = set()
    for date_obj in available_months:
        month_abbr = date_obj.strftime("%b")
        year = date_obj.year
        valid_months.add(f"{month_abbr} {year}")

    # Count each status by month
    status_data = (
        Membership.objects
        .annotate(mos_date=Cast('mos', DateField()))  # Ensure correct date conversion
        .filter(mos_date__gte=twelve_months_ago, stat__isnull=False)
        .exclude(stat="")  # Exclude empty states
        .annotate(year=TruncYear('mos_date'), month=TruncMonth('mos_date'))
        .values('year', 'month', 'stat')
        .annotate(count=Count('member_id'))
        .order_by('-year', '-month')
    )

    # Structure data correctly without unnecessary mappings
    status_dict = {}
    for entry in status_data:
        year = entry['year'].year
        month_abbr = entry['month'].strftime("%b")
        month_key = f"{month_abbr} {year}"  # Ex: "Jan 2025"

        if month_key not in valid_months:
            continue  # Only include months that exist in the database

        if month_key not in status_dict:
            status_dict[month_key] = {"ENROLLED": 0, "REENROLLED": 0, "DISENROLLED": 0}

        # Use 'stat' value directly
        raw_stat = entry['stat'].strip().upper()

        if raw_stat in status_dict[month_key]:
            status_dict[month_key][raw_stat] += entry['count']

    # Guardar en caché por 1 hora
    cache.set(cache_key, status_dict, 3600)

    return JsonResponse(status_dict)


# Enhanced view for status report
@login_required
def status_all_plans_view(request):
    """
    Enhanced view for status report across all plans with period filtering.
    """
    year_param = request.GET.get("year", "last_12")
    user_id = getattr(request.user, 'id', 'anonymous')

    # Optimized cache key
    cache_key = f'status_all_plans_v2_{year_param}_{user_id}'
    cached_context = cache.get(cache_key)

    if cached_context:
        # Crear datos del breadcrumb y agregarlos al contexto en caché
        period_text = "Last 12 Months" if cached_context['selected_year'] == "last_12" else cached_context[
            'selected_year']
        breadcrumb_data = [
            {"label": "Dashboard", "url": reverse("home")},
            {"label": f"Status Reports - {period_text}"}
        ]
        cached_context['breadcrumb_data'] = breadcrumb_data

        request.session['yearly_report_data'] = cached_context['pivot_list']
        request.session['yearly_field_names'] = ['plan', 'stat'] + cached_context['capmo_labels']
        return render(request, 'webapp/status_all_plans.html', cached_context)

    today = now().date().replace(day=1)
    twelve_months_ago = today - relativedelta(months=11)

    # Determine the date filter based on selected period
    if year_param == "last_12":
        selected_year = "last_12"
        start_date = twelve_months_ago
        end_date = today
        month_range = [
            (start_date + relativedelta(months=i)).strftime("%b-%Y")
            for i in range(12)
        ]
    else:
        try:
            selected_year = int(year_param)
            start_date = datetime(selected_year, 1, 1).date()
            end_date = datetime(selected_year, 12, 31).date()
            month_range = [
                datetime(selected_year, month, 1).strftime("%b-%Y")
                for month in range(1, 13)
            ]
        except ValueError:
            selected_year = "last_12"
            start_date = twelve_months_ago
            end_date = today
            month_range = [
                (start_date + relativedelta(months=i)).strftime("%b-%Y")
                for i in range(12)
            ]

    # Adjust filter according to the selected period
    if selected_year == "last_12":
        date_filter = {
            'mos__gte': start_date.strftime("%Y-%m-01"),
            'mos__lte': end_date.replace(day=28).strftime("%Y-%m-%d")
        }
    else:
        date_filter = {
            'mos__startswith': str(selected_year)
        }

    # Fetch status data with the appropriate filter
    pivot_data = (
        Membership.objects
        .filter(**date_filter)
        .exclude(stat__isnull=True)
        .exclude(stat__in=['', 'None'])
        .values('plan', 'stat', 'mos')
        .annotate(total=Count('member_id'))
        .order_by('plan', 'stat', 'mos')
    )

    pivot_dict = {}
    total_by_month = {}

    for row in pivot_data:
        # Convert to month-year format (e.g., "Jan-2025")
        try:
            capmo_str = datetime.strptime(str(row['mos']), "%Y-%m-%d").strftime("%b-%Y")
        except ValueError:
            # Handle case where 'mos' might already be in YYYY-MM format
            year_month = str(row['mos']).split('-')
            if len(year_month) >= 2:
                try:
                    month_num = int(year_month[1])
                    year_num = int(year_month[0])
                    month_name = calendar.month_abbr[month_num]
                    capmo_str = f"{month_name}-{year_num}"
                except (ValueError, IndexError):
                    continue
            else:
                continue

        # Only process if this month is in our selected range
        if capmo_str in month_range:
            # Update totals
            total_by_month[capmo_str] = total_by_month.get(capmo_str, 0) + row['total']

            # Update pivot dictionary
            key = (row['plan'], row['stat'])
            if key not in pivot_dict:
                pivot_dict[key] = {}
            pivot_dict[key][capmo_str] = row['total']

    # Sort months chronologically
    sorted_capmo_labels = sorted(month_range, key=lambda x: datetime.strptime(x, "%b-%Y"))

    # Create the pivot list with all months in range, filling in zeros for missing data
    pivot_list = [
        {"plan": plan, "stat": stat,
         **{capmo: pivot_dict.get((plan, stat), {}).get(capmo, 0) for capmo in sorted_capmo_labels}}
        for (plan, stat) in pivot_dict.keys()
    ]

    # Save data for export
    request.session['yearly_report_data'] = pivot_list
    request.session['yearly_field_names'] = ['plan', 'stat'] + sorted_capmo_labels

    # Get available years for the dropdown
    available_years = Membership.objects.annotate(
        mos_date=Cast('mos', DateField())
    ).dates('mos_date', 'year', order='DESC')

    # Crear datos del breadcrumb
    period_text = "Last 12 Months" if selected_year == "last_12" else selected_year
    breadcrumb_data = [
        {"label": "Dashboard", "url": reverse("home")},
        {"label": f"Status Reports - {period_text}"}
    ]

    # Prepare the context for the template
    context = {
        'pivot_list': pivot_list,
        'capmo_labels': sorted_capmo_labels,
        'total_by_month': total_by_month,
        'selected_year': selected_year,
        'available_years': available_years,
        'report_model': 'Membership',
        'breadcrumb_data': breadcrumb_data,  # Añadir datos del breadcrumb al contexto
    }

    # Cache for 24 hours
    cache.set(cache_key, context, 3600 * 24)

    return render(request, 'webapp/status_all_plans.html', context)


@login_required
def status_by_plans_view(request):
    """
    View for the status report broken down by plan
    """
    # Prepare the context to send to the template
    context = {
        'report_model': 'Membership',
        'report_title': 'Member Status by Plan'
    }

    return render(request, 'webapp/status_by_plans.html', context)


# ✅ Export Detail Report to Excel
# Función modificada para manejar filtros de status
def export_detail_to_excel(request, plan, capmo):
    """
    Exports the detail data to an Excel file, with full-table search support.
    """
    try:
        capmo_date = datetime.strptime(capmo, "%b-%Y").strftime("%Y-%m")
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    stat = request.resolver_match.kwargs.get('stat')
    search = request.GET.get('search', '').strip().lower()

    # Base queryset
    if stat:
        if plan == 'TOTAL':
            base_query = Membership.objects.filter(mos__startswith=capmo_date, mshp=1, stat=stat)
        else:
            base_query = Membership.objects.filter(plan=plan, mos__startswith=capmo_date, mshp=1, stat=stat)
    else:
        if plan == 'TOTAL':
            base_query = Membership.objects.filter(mos__startswith=capmo_date, mshp=1)
        else:
            base_query = Membership.objects.filter(plan=plan, mos__startswith=capmo_date, mshp=1)

    # Convert to list of dicts
    data = list(base_query.values())

    # Fields used in export
    field_names = [
        "center", "plan", "lob", "member_id", "medicare_id", "medicaid_id",
        "member_name", "dob", "age", "sex", "address", "city", "st", "zip",
        "county", "phone_number", "mos", "pcp_name"
    ]
    if stat:
        field_names.append("stat")

    # Filter by search string across all fields
    if search:
        def matches(row):
            for field in field_names:
                value = row.get(field)
                if value is not None and search in str(value).lower():
                    return True
            return False

        data = list(filter(matches, data))

    if not data:
        return HttpResponse("No data available to export", status=400)

    # Prepare Excel response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{plan}_{capmo}"
    if stat:
        filename = f"{plan}_{stat}_{capmo}"
    response['Content-Disposition'] = f'attachment; filename="{filename}_detail_report.xlsx"'

    # Create workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Detail Report"

    # Headers
    for col_num, column_title in enumerate(field_names, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for row_num, item in enumerate(data, start=2):
        for col_num, field in enumerate(field_names, 1):
            value = item.get(field, '')
            if field == 'dob' and isinstance(value, date):
                value = value.strftime('%Y-%m-%d')
            if field == 'mos' and isinstance(value, date):
                value = value.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=col_num).value = value

    workbook.save(response)
    return response


@login_required
def cap_detail_status_view(request, stat, capmo):
    """
    View to show the detail of the status report.
    """
    try:
        # Try to parse standard "Jan-2025" format
        capmo_date = datetime.strptime(capmo, "%b-%Y").strftime("%Y-%m")
    except ValueError:
        try:
            # Try alternative format
            capmo_date = datetime.strptime(capmo, "%B-%Y").strftime("%Y-%m")
        except ValueError:
            try:
                # Third attempt: maybe it's already in "2025-01" format
                capmo_date = datetime.strptime(capmo, "%Y-%m").strftime("%Y-%m")
            except ValueError:
                return HttpResponse(f"Invalid date format: {capmo}", status=400)

    # Plan comes as an optional parameter in the querystring
    plan_filter = request.GET.get('plan', None)

    # Build filter conditions
    filters = {'stat': stat, 'mos__startswith': capmo_date}

    # Add plan filter if provided
    if plan_filter and plan_filter != 'TOTAL':
        filters['plan'] = plan_filter

    # Filter by STAT, MOS and optionally by PLAN
    data = Membership.objects.filter(**filters)

    if not data.exists():
        return HttpResponse(f"No data available for the selected period ({capmo}) with status {stat}.", status=404)

    # Convert QuerySet to list for modification
    data_list = list(data.values())

    # Convert date objects to strings for serialization
    for item in data_list:
        if 'dob' in item and isinstance(item['dob'], date):
            item['dob'] = item['dob'].strftime('%Y-%m-%d')
        # Convert any other date fields
        if 'mos' in item and isinstance(item['mos'], date):
            item['mos'] = item['mos'].strftime('%Y-%m-%d')

    # Save data in session for export
    request.session['detail_report_data'] = data_list
    request.session['detail_field_names'] = [
        "center", "plan", "lob", "mshp", "member_id", "medicare_id", "medicaid_id",
        "member_name", "dob", "age", "sex", "address", "city", "st", "zip",
        "county", "phonenumber", "mos", "pcpname", "stat"  # Added stat field
    ]

    # Obtener parámetros de la URL
    origin = request.GET.get('origin', 'status_all_plans')
    year_param = request.GET.get('year', 'last_12')

    # Determinar etiqueta para el breadcrumb
    year_label = year_param
    if year_param == 'last_12':
        year_label = 'Last 12 Months'

    # Construir el texto del último breadcrumb usando plan o stat
    breadcrumb_label = f"{plan_filter or stat} ({capmo})"

    # Breadcrumb para la plantilla
    breadcrumb_data = [
        {"label": "Dashboard", "url": reverse("home")},
        {"label": f"Status Reports - {year_label}",
         "url": f"{reverse('status_all_plans')}?year={year_param}"},
        {"label": breadcrumb_label}
    ]

    # Armado del contexto (incluir year_param)
    context = {
        'data': data,
        'plan': plan_filter or stat,
        'capmo': capmo,
        'origin': origin,
        'year_param': year_param,  # Añadir al contexto para uso en template
        'report_model': 'Membership',
        'breadcrumb_data': breadcrumb_data,
    }

    return render(request, 'webapp/cap_detail_status.html', context)


# Revenue Report View
class RevenueReportView(TemplateView):
    template_name = "webapp/revenue_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        year_param = self.request.GET.get("year", "last_12")

        # Clave de caché optimizada por usuario
        user_id = getattr(self.request.user, 'id', 'anonymous')
        cache_key = f'revenue_data_{year_param}_{user_id}'

        cached_data = cache.get(cache_key)
        if cached_data:
            context.update(cached_data)

            # ✅ Breadcrumb también cuando usamos caché
            selected_year_display = "Last 12 Months" if str(year_param) == "last_12" else str(year_param)
            breadcrumb_data = [
                {"label": "Dashboard", "url": reverse("home")},
                {"label": f"Financial Reports - {selected_year_display}"}
            ]
            context["breadcrumb_data"] = breadcrumb_data
            context["selected_year"] = str(year_param)

            return context

        today = now().date().replace(day=1)
        twelve_months_ago = today - relativedelta(months=11)

        if year_param == "last_12":
            selected_year = "last_12"
            start_dates = [
                (twelve_months_ago + relativedelta(months=i)).strftime("%Y-%m")
                for i in range(12)
            ]
            month_range = [
                (twelve_months_ago + relativedelta(months=i)).strftime("%b-%Y")
                for i in range(12)
            ]
        else:
            try:
                selected_year = int(year_param)
                month_range = [datetime(selected_year, month, 1).strftime("%b-%Y") for month in range(1, 13)]
                start_dates = [f"{selected_year}-{month:02d}" for month in range(1, 13)]
            except ValueError:
                selected_year = today.year
                month_range = [datetime(today.year, month, 1).strftime("%b-%Y") for month in range(1, 13)]
                start_dates = [f"{today.year}-{month:02d}" for month in range(1, 13)]

        month_format_map = {}
        for month_str in month_range:
            dt = datetime.strptime(month_str, "%b-%Y")
            db_format = dt.strftime("%Y-%m")
            month_format_map[db_format] = month_str

        pivot_dict = {}
        total_by_month = {month: 0 for month in month_range}
        placeholders = ', '.join(['%s'] * len(start_dates))

        with connection.cursor() as cursor:
            query = f"""
                SELECT m.MemberFullName, m.MedicareId,
                       SUBSTRING(m.MOS, 1, 7) as month_year,
                       SUM(m.ProviderFundBalance) as total_balance,
                       CASE WHEN cl.MedicareId IS NOT NULL THEN 1 ELSE 0 END as has_claims
                FROM providerlineal m
                LEFT JOIN (
                    SELECT DISTINCT MedicareId, SUBSTRING(MOS, 1, 7) as month_year
                    FROM claimlineal
                    WHERE SUBSTRING(MOS, 1, 7) IN ({placeholders})
                ) cl ON m.MedicareId = cl.MedicareId AND SUBSTRING(m.MOS, 1, 7) = cl.month_year
                WHERE SUBSTRING(m.MOS, 1, 7) IN ({placeholders})
                GROUP BY m.MemberFullName, m.MedicareId, SUBSTRING(m.MOS, 1, 7), has_claims
                ORDER BY m.MemberFullName, month_year
            """
            cursor.execute(query, start_dates + start_dates)
            revenue_data = cursor.fetchall()

            cursor.execute("""
                SELECT DISTINCT SUBSTRING(`MOS`, 1, 4) as year 
                FROM providerlineal 
                ORDER BY year DESC
            """)
            available_years = [{'year': row[0]} for row in cursor.fetchall()]

        for row in revenue_data:
            member_name = row[0]
            medicare_id = row[1]
            month_year_db = row[2]
            balance = float(row[3])
            has_claims = bool(row[4])

            if month_year_db in month_format_map:
                capmo_str = month_format_map[month_year_db]

                if member_name not in pivot_dict:
                    pivot_dict[member_name] = {
                        'medicare_id': medicare_id,
                        **{month: 0 for month in month_range},
                        **{f"{month}_has_claims": False for month in month_range}
                    }

                pivot_dict[member_name][capmo_str] = balance
                pivot_dict[member_name][f"{capmo_str}_has_claims"] = has_claims
                total_by_month[capmo_str] += balance

        pivot_list = []
        for member, data in pivot_dict.items():
            formatted_name = ' '.join(word.lower().capitalize() for word in member.split())
            pivot_list.append({
                "member": formatted_name,
                "medicare_id": data['medicare_id'],
                **{k: v for k, v in data.items() if k != 'medicare_id'}
            })

        pivot_list.sort(key=lambda x: x['member'])

        cacheable_data = {
            'pivot_list': pivot_list,
            'capmo_labels': month_range,
            'month_range': month_range,
            'total_by_month': total_by_month,
            'selected_year': str(selected_year),
            'available_years': available_years,
            'report_model': 'ProviderLineal',
        }

        cache.set(cache_key, cacheable_data, 3600)
        context.update(cacheable_data)

        # ✅ Breadcrumb fuera del caché
        selected_year_display = "Last 12 Months" if str(selected_year) == "last_12" else str(selected_year)
        breadcrumb_data = [
            {"label": "Dashboard", "url": reverse("home")},
            {"label": f"Financial Reports - {selected_year_display}"}
        ]
        context["breadcrumb_data"] = breadcrumb_data

        # ✅ 🔥 GUARDA los datos para el botón de exportación
        self.request.session['revenue_report_data'] = cacheable_data['pivot_list']
        self.request.session['revenue_report_field_names'] = cacheable_data['capmo_labels']

        return context


# Export revenue report to Excel
def export_revenue_to_excel(request):
    """
    Exports revenue report data to an Excel file using data saved in session.
    Applies optional filtering based on search input.
    """

    # Retrieve data and column headers from session
    pivot_list = request.session.get('revenue_report_data', [])
    field_names = request.session.get('revenue_report_field_names', [])

    # Apply optional search filter
    search = request.GET.get("search", "").strip().lower()
    if search and pivot_list:
        pivot_list = [
            row for row in pivot_list
            if any(search in str(value).lower() for value in row.values())
        ]

    # Validate we have data and headers
    if not pivot_list or not field_names:
        return HttpResponse("No data available to export", status=400)

    # Create a pandas DataFrame using the headers (for consistent column order)
    df = pd.DataFrame(pivot_list, columns=field_names)

    # Prepare HTTP response for Excel download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="revenue_report.xlsx"'

    # Write the DataFrame to Excel using openpyxl engine
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Revenue Report")

    return response



# API endpoint for financial data
def get_financial_data(request):
    """
    Returns financial data in JSON format for dashboard charts.
    """
    cache_key = 'financial_data'
    cached_data = cache.get(cache_key)

    if cached_data:
        return JsonResponse(cached_data)

    today = now().date().replace(day=1)
    twelve_months_ago = today - relativedelta(months=11)

    financial_data = (
        ProviderLineal.objects
        .annotate(mos_date=Cast('mos', DateField()))
        .filter(mos_date__gte=twelve_months_ago)
        .annotate(year=TruncYear('mos_date'), month=TruncMonth('mos_date'))
        .values('year', 'month')
        .annotate(total_balance=Sum('provider_fund_balance'))
        .order_by('-year', '-month')
    )

    data = {}
    for entry in financial_data:
        year = entry['year'].year
        month_abbr = entry['month'].strftime("%b")
        balance = float(entry['total_balance'])
        date_key = f"{month_abbr} {year}"
        data[date_key] = {'total': balance}

    sorted_data = dict(sorted(data.items(),
                              key=lambda x: datetime.strptime(x[0], "%b %Y"),
                              reverse=True))

    # Guardar en caché por 1 hora
    cache.set(cache_key, sorted_data, 3600)

    return JsonResponse(sorted_data)


@login_required
def claim_detail_view(request, medicare_id, mos):
    cache_key = f'claim_detail_v2_{medicare_id}_{mos}'
    cached_data = cache.get(cache_key)

    selected_year = request.GET.get("year", "last_12")

    # ✅ If cached, inject breadcrumb dynamically before returning
    if cached_data:
        member_name = cached_data.get("member_name", "Unknown")
        period_label = "Last 12 Months" if selected_year == "last_12" else selected_year

        cached_data["breadcrumb_data"] = [
            {"label": "Dashboard", "url": reverse("home")},
            {"label": f"Financial Reports - {period_label}", "url": request.META.get("HTTP_REFERER", reverse("revenue_report"))},
            {"label": f"Claims for {escape(member_name)} ({mos})"}
        ]

        return render(request, 'webapp/claim_detail.html', cached_data)

    try:
        # 🔄 Normalize month
        if '-' in mos:
            month_name, year = mos.split('-')
            try:
                month_date = datetime.strptime(f"{month_name} {year}", "%b %Y")
            except ValueError:
                month_date = datetime.strptime(f"{month_name} {year}", "%B %Y")
            mos_date = month_date.strftime("%Y-%m")
        else:
            mos_date = mos

        export_fields = [
            'MOS', 'ClaimId', 'ClaimLine', 'MedicareId', 'MemFullName', 'POS',
            'ClaimStartDate', 'ClaimEndDate', 'PaidDate', 'ClaimDetailStatus',
            'AmountPaid', 'AdminFee', 'ProvFullName', 'ProvSpecialty', 'MemDOB',
            'MemAge', 'MemPCPFullName', 'CarrierMemberID', 'MemEnrollId',
            'Diagnoses', 'AllowAmt', 'PharmacyName', 'NPOS', 'Pharmacy',
            'Claims', 'County_Simple', 'NPOS_Simple', 'Triangle_Cover'
        ]

        raw_queryset = ClaimLineal.objects.filter(
            MedicareId=medicare_id,
            MOS__startswith=mos_date
        ).values(*export_fields).order_by('ClaimId', 'ClaimLine')

        data_list = []
        for item in raw_queryset:
            for key, value in item.items():
                if isinstance(value, date):
                    item[key] = value.strftime('%Y-%m-%d')
                elif isinstance(value, (datetime, time)):
                    item[key] = str(value)
            data_list.append(item)

        if not data_list:
            return HttpResponse(f"No data found for Medicare ID {medicare_id} in {mos}.", status=404)

        member_name = data_list[0].get("MemFullName", "Unknown")
        period_label = "Last 12 Months" if selected_year == "last_12" else selected_year

        # ✅ Full breadcrumb
        breadcrumb_data = [
            {"label": "Dashboard", "url": reverse("home")},
            {"label": f"Financial Reports - {period_label}", "url": request.META.get("HTTP_REFERER", reverse("revenue_report"))},
            {"label": f"Claims for {escape(member_name)} ({mos})"}
        ]

        context = {
            'data': data_list,
            'medicare_id': medicare_id,
            'mos': mos,
            'member_name': member_name,
            'report_model': 'ClaimLineal',
            'breadcrumb_data': breadcrumb_data  # ✅ breadcrumb is now always present
        }

        request.session['claim_detail_data'] = data_list
        request.session['claim_field_names'] = export_fields

        # ✅ Cache context WITHOUT breadcrumb to avoid storing outdated links
        cache.set(cache_key, {
            'data': data_list,
            'medicare_id': medicare_id,
            'mos': mos,
            'member_name': member_name,
            'report_model': 'ClaimLineal'
        }, 3600 * 24)

        return render(request, 'webapp/claim_detail.html', context)

    except Exception as e:
        return HttpResponse(f"Error processing request: {str(e)}", status=500)


@login_required
def export_claims_excel(request, medicare_id, mos):
    """
    Exports claim data to an Excel file.
    """
    # Get data from session
    claim_data = request.session.get('claim_detail_data', [])
    field_names = request.session.get('claim_field_names', [])

    if not claim_data:
        return HttpResponse("No data available to export", status=400)

    # Convert to DataFrame
    df = pd.DataFrame(claim_data, columns=field_names)

    # Create HTTP response with Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="claims_{medicare_id}_{mos}.xlsx"'

    # Write to Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Claims Detail")

    return response


def export_status_detail(request, stat, capmo):
    """
    Export detailed status data to Excel, with support for search and filtering
    """
    # Attempt to convert 'capmo' into a standard YYYY-MM format
    try:
        capmo_date = datetime.strptime(capmo, "%b-%Y").strftime("%Y-%m")
    except ValueError:
        try:
            capmo_date = datetime.strptime(capmo, "%B-%Y").strftime("%Y-%m")
        except ValueError:
            try:
                capmo_date = datetime.strptime(capmo, "%Y-%m").strftime("%Y-%m")
            except ValueError:
                return HttpResponse(f"Invalid date format: {capmo}", status=400)

    # Get filters from GET parameters
    plan_filter = request.GET.get('plan', None)
    search_text = request.GET.get('search', None)

    # Normalize 'stat' and handle known aliases
    stat_upper = stat.upper().strip()
    stat_map = {
        "ENROLLED": ["ENROLLED", "NEW"],
        "REENROLLED": ["REENROLLED", "REENROL"],
        "DISENROLLED": ["DISENROLLED", "TERM", "TERMINATED"]
    }
    valid_stats = stat_map.get(stat_upper, [stat])

    # Determine mshp value based on status
    mshp_value = 1
    if stat_upper in ["DISENROLLED", "TERM", "TERMINATED"]:
        mshp_value = 0

    # Build query filters
    filters = {
        'stat__in': valid_stats,
        'mos__startswith': capmo_date,
        'mshp': mshp_value
    }

    if plan_filter and plan_filter != 'TOTAL':
        filters['plan'] = plan_filter

    # Execute query with base filters
    query = Membership.objects.filter(**filters)

    # Apply search if provided
    if search_text and search_text.strip():
        search_query = Q()
        search_fields = [
            'center', 'plan', 'lob', 'member_id', 'medicare_id',
            'medicaid_id', 'member_name', 'address', 'city', 'st',
            'zip', 'county', 'phonenumber', 'pcpname'
        ]
        for field in search_fields:
            search_query |= Q(**{f"{field}__icontains": search_text})
        query = query.filter(search_query)

    # If no data found, return early
    if not query.exists():
        return HttpResponse(f"No data available for {stat} in {capmo}", status=400)

    # Prepare Excel file response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Status_{stat}_{capmo}"
    if plan_filter:
        filename += f"_{plan_filter}"
    if search_text:
        filename += "_filtered"
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    # Define column fields
    field_names = [
        "center", "plan", "lob", "member_id", "medicare_id", "medicaid_id",
        "member_name", "dob", "age", "sex", "address", "city", "st", "zip",
        "county", "phone_number", "mos", "pcp_name", "stat"
    ]

    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"Status {stat}"

    # Write headers
    for col_num, column_title in enumerate(field_names, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = openpyxl.styles.Font(bold=True)

    # Write data rows
    row_num = 2
    for item in query.values():
        for col_num, field in enumerate(field_names, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            value = item.get(field, '')

            # Format date fields
            if field == 'dob' and isinstance(value, date):
                value = value.strftime('%Y-%m-%d')
            if field == 'mos' and isinstance(value, date):
                value = value.strftime('%Y-%m-%d')

            cell.value = value
        row_num += 1

    workbook.save(response)
    return response
